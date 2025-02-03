#!C:\VScode\venv\Scripts\python.exe
# -*- coding: utf-8 -*-

import os
from pathlib import Path
from colorama import init, Fore, Style
from openpyxl import Workbook, load_workbook

# Инициализация colorama для поддержки раскраски в терминале
init(autoreset=True)

# Хранилище для сохранения текстовых значений
results_storage = []

def delete_old_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"Старый файл удален: {file_path}")
        
def check_keyword_in_file(file_path, keyword, success_message, failure_message, invert_logic=False, required_count=None):
    # Проверка существования файла
    if not os.path.isfile(file_path):
        result = f"Файл не найден: {file_path}"
        status = "✗"
    else:
        # Поиск ключевого слова в содержимом файла
        with open(file_path, "r", encoding="utf-8") as file:
            file_content = file.read()
            keyword_count = file_content.count(keyword)  # Подсчёт вхождений ключевого слова
            
            # Если указано required_count, проверяем количество вхождений
            if required_count is not None:
                condition = keyword_count == required_count  # Проверяем на нужное количество вхождений
            else:
                condition = keyword in file_content  # Просто проверка на наличие ключевого слова
            
            # Если инвертировать логику
            if invert_logic:
                condition = not condition
            
            # Определяем результат в зависимости от выполнения условия
            if condition:
                result = success_message
                status = "✓"
            else:
                result = failure_message
                status = "✗"

    # Вывод в терминал
    print(result)
    
    # Сохранение результата в хранилище
    results_storage.append((status, result))


if __name__ == "__main__":
    # Проверка МодификацияКонфигурацииПереопределяемый
    xml_file_path_1 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\МодификацияКонфигурацииПереопределяемый\Ext\Module.bsl"
    keyword_module = "АЗС_СобытияФорм.ПриСозданииНаСервере(Форма, Отказ, СтандартнаяОбработка)"
    check_keyword_in_file(
        xml_file_path_1,
        keyword_module,
        "✓ МодификацияКонфигурацииПереопределяемый Добавлен код ПередЗаписьюНаСервере(Форма, Отказ, ТекущийОбъект, ПараметрыЗаписи)Экспорт.",
        "✗ МодификацияКонфигурацииПереопределяемый не добавлен код ПередЗаписьюНаСервере(Форма, Отказ, ТекущийОбъект, ПараметрыЗаписи)Экспорт."
    )


    # Проверка СвязанныеДокументы
    xml_file_path_2 = r"E:\Users\AUshanov\izmenenie_tip\CommonCommands\СвязанныеДокументы.xml"
    keyword_module = "cfg:DocumentRef.АЗС_ПоступлениеТоплива"
    check_keyword_in_file(
        xml_file_path_2,
        keyword_module,
        "✓ В тип параметра команды СвязанныеДокументы добавлен ДокументСсылка.АЗС_ПоступлениеТоплива.",
        "✗ В тип параметра команды СвязанныеДокументы не добавлен ДокументСсылка.АЗС_ПоступлениеТоплива."
    )


    # Проверка СвязанныеДокументы
    xml_file_path_3 = r"E:\Users\AUshanov\izmenenie_tip\CommonCommands\СвязанныеДокументы.xml"
    keyword_module = "cfg:DocumentRef.АЗС_СменныйОтчет"
    check_keyword_in_file(
        xml_file_path_3,
        keyword_module,
        "✓ В тип параметра команды СвязанныеДокументы добавлен ДокументСсылка.АЗС_СменныйОтчет.",
        "✗ В тип параметра команды СвязанныеДокументы не добавлен ДокументСсылка.АЗС_СменныйОтчет."
    )


    # Проверка СвязанныеДокументы
    xml_file_path_4 = r"E:\Users\AUshanov\izmenenie_tip\CommonCommands\СвязанныеДокументы.xml"
    keyword_module = "cfg:DocumentRef.УдалитьАЗС_СборкаТоваров"
    check_keyword_in_file(
        xml_file_path_4,
        keyword_module,
        "✓ В тип параметра команды СвязанныеДокументы добавлен ДокументСсылка.УдалитьАЗС_СборкаТоваров.",
        "✗ В тип параметра команды СвязанныеДокументы не добавлен ДокументСсылка.УдалитьАЗС_СборкаТоваров."
    )
    
       # Проверка ВариантыОтчетовПереопределяемый
    xml_file_path_5 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ВариантыОтчетовПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_5,
        keyword_module,
        "✓ В ОМ ВариантыОтчетовПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ ВариантыОтчетовПереопределяемый  не добавлен ПРАКТИКОН."
    )
     
       # Проверка ВариантыОтчетовУТПереопределяемый
    xml_file_path_6 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ВариантыОтчетовУТПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_6,
        keyword_module,
        "✓ В ОМ ВариантыОтчетовУТПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ ВариантыОтчетовУТПереопределяемый  не добавлен ПРАКТИКОН.",
        invert_logic=True
    )
       # Проверка ЗакупкиКлиент
    xml_file_path_7 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ЗакупкиКлиент\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_7,
        keyword_module,
        "✓ В ОМ ЗакупкиКлиент  добавлен ПРАКТИКОН",
        "✗ В ОМ ЗакупкиКлиент  не добавлен ПРАКТИКОН."
    )

    # Проверка ЗапретРедактированияРеквизитовОбъектовПереопределяемый
    xml_file_path_8 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ЗапретРедактированияРеквизитовОбъектовПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_8,
        keyword_module,
        "✓ В ОМ ЗапретРедактированияРеквизитовОбъектовПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ ЗапретРедактированияРеквизитовОбъектовПереопределяемый  не добавлен ПРАКТИКОН."
    )
    # Проверка МенеджерОбменаЧерезУниверсальныйФормат
    xml_file_path_9 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\МенеджерОбменаЧерезУниверсальныйФормат\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_9,
        keyword_module,
        "✓ В ОМ МенеджерОбменаЧерезУниверсальныйФормат  добавлен ПРАКТИКОН",
        "✗ В ОМ МенеджерОбменаЧерезУниверсальныйФормат  не добавлен ПРАКТИКОН."
    )
     # Проверка МодификацияКонфигурацииКлиентПереопределяемый
    xml_file_path_10 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\МодификацияКонфигурацииКлиентПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_10,
        keyword_module,
        "✓ В ОМ МодификацияКонфигурацииКлиентПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ МодификацияКонфигурацииКлиентПереопределяемый  не добавлен ПРАКТИКОН."
    )
      # Проверка МодификацияКонфигурацииПереопределяемый
    xml_file_path_11 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\МодификацияКонфигурацииПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_11,
        keyword_module,
        "✓ В ОМ МодификацияКонфигурацииПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ МодификацияКонфигурацииПереопределяемый  не добавлен ПРАКТИКОН."
    )
      # Проверка ПодсистемыКонфигурацииПереопределяемый
    xml_file_path_12 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ПодсистемыКонфигурацииПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_12,
        keyword_module,
        "✓ В ОМ ПодсистемыКонфигурацииПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ ПодсистемыКонфигурацииПереопределяемый  не добавлен ПРАКТИКОН."
    )
     # Проверка ПолучениеОбновленийПрограммыПереопределяемый (приведен к типовому)
    xml_file_path_13 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ПолучениеОбновленийПрограммыПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
    xml_file_path_13,
    keyword_module,
    "✓ В ОМ ПолучениеОбновленийПрограммыПереопределяемый (приведен к типовому)  не добавлен ПРАКТИКОН",
    "✗ В ОМ ПолучениеОбновленийПрограммыПереопределяемый (приведен к типовому)  добавлен ПРАКТИКОН.",
    invert_logic=True
)

    # Проверка ПроведениеДокументовЛокализация
    xml_file_path_14 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ПроведениеДокументовЛокализация\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_14,
        keyword_module,
        "✓ В ОМ ПроведениеДокументовЛокализация  добавлен ПРАКТИКОН",
        "✗ В ОМ ПроведениеДокументовЛокализация  не добавлен ПРАКТИКОН."
    )
      # Проверка СкидкиНаценкиСервер (приведен к типовому) (приведен к типовому)
    xml_file_path_15 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\СкидкиНаценкиСервер\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_15,
        keyword_module,
        "✓ В ОМ СкидкиНаценкиСервер (приведен к типовому)  не добавлен ПРАКТИКОН",
        "✗ В ОМ СкидкиНаценкиСервер (приведен к типовому)  добавлен ПРАКТИКОН.",
        invert_logic=True
    )
        # Проверка УправлениеДоступомПереопределяемый
    xml_file_path_16 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\УправлениеДоступомПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_16,
        keyword_module,
        "✓ В ОМ УправлениеДоступомПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ УправлениеДоступомПереопределяемый  не добавлен ПРАКТИКОН.",
        required_count=4
    )

    # Проверка УправлениеПечатьюПереопределяемый
    xml_file_path_17 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\УправлениеПечатьюПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_17,
        keyword_module,
        "✓ В ОМ УправлениеПечатьюПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ УправлениеПечатьюПереопределяемый  не добавлен ПРАКТИКОН."
    )
    # Проверка УправлениеСвойствами
    xml_file_path_18 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\УправлениеСвойствами\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_18,
        keyword_module,
        "✓ В ОМ УправлениеСвойствами  добавлен ПРАКТИКОН",
        "✗ В ОМ УправлениеСвойствами  не добавлен ПРАКТИКОН."
    )
     # Проверка УчетНДСУПККлиентСервер
    xml_file_path_19 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\УчетНДСУПКлиентСервер\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_19,
        keyword_module,
        "✓ В ОМ УчетНДСУПККлиентСервер  добавлен ПРАКТИКОН",
        "✗ В ОМ УчетНДСУПККлиентСервер  не добавлен ПРАКТИКОН."
    )
      # Проверка Ценообразование
    xml_file_path_20 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\Ценообразование\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_20,
        keyword_module,
        "✓ В ОМ Ценообразование  добавлен ПРАКТИКОН",
        "✗ В ОМ Ценообразование  не добавлен ПРАКТИКОН."
    )
      # Проверка ТекущиеДелаПереопределяемый
    xml_file_path_21 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ТекущиеДелаПереопределяемый\Ext\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_21,
        keyword_module,
        "✓ В ОМ ТекущиеДелаПереопределяемый  добавлен ПРАКТИКОН",
        "✗ В ОМ ТекущиеДелаПереопределяемый  не добавлен ПРАКТИКОН."
    )


      # Проверка БизнесРегионы
    xml_file_path_22 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\БизнесРегионы.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_22,
        keyword_module,
        "✓ В ОМ БизнесРегионы  добавлены  реквизиты АЗС",
        "✗ В ОМ БизнесРегионы  не добавлены  реквизитов АЗС."
    )
      # Проверка ВидыКонтактнойИнформации
    xml_file_path_23 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ВидыКонтактнойИнформации.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_23,
        keyword_module,
        "✓ В Справочнике ВидыКонтактнойИнформации   Изменен состав",
        "✗ В Справочнике ВидыКонтактнойИнформации  не Изменен Изменен состав."
    )
  # Проверка ВидыНоменклатуры
    xml_file_path_24 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ВидыНоменклатуры.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_24,
        keyword_module,
        "✓ В Справочнике ВидыНоменклатуры   Изменен состав",
        "✗ В Справочнике ВидыНоменклатуры  не Изменен Изменен состав."
    )
   # Проверка ВидыЦен
    xml_file_path_25 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ВидыЦен.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_25,
        keyword_module,
        "✓ В Справочнике ВидыЦен   Изменен состав",
        "✗ В Справочнике ВидыЦен  не Изменен Изменен состав."
    )   
     # Проверка ДоговорыКонтрагентов
    xml_file_path_26 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ДоговорыКонтрагентов.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_26,
        keyword_module,
        "✓ В Справочнике ДоговорыКонтрагентов  Изменен состав",
        "✗ В Справочнике ДоговорыКонтрагентов  не Изменен Изменен состав."
    ) 
      # Проверка ДоговорыКонтрагентов
    xml_file_path_27 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ДоговорыКонтрагентов\Forms\ФормаЭлемента\Ext\Form\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_27,
        keyword_module,
        "✓ В Справочнике ДоговорыКонтрагентов  Изменена форма элемента",
        "✗ В Справочнике ДоговорыКонтрагентов  не Изменена форма элемента."
    ) 
       # Проверка ЗначенияСвойствОбъектов
    xml_file_path_28 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ЗначенияСвойствОбъектов.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_28,
        keyword_module,
        "✓ В Справочнике ЗначенияСвойствОбъектов  Изменен состав",
        "✗ В Справочнике ЗначенияСвойствОбъектов  не Изменен состав."
    )
        # Проверка Кассы
    xml_file_path_29 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Кассы.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_29,
        keyword_module,
        "✓ В Справочнике Кассы  Изменен состав",
        "✗ В Справочнике Кассы  не Изменен состав."
    )
         # Проверка КассыККМ
    xml_file_path_30 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\КассыККМ.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_30,
        keyword_module,
        "✓ В Справочнике КассыККМ  Изменен состав",
        "✗ В Справочнике КассыККМ  не Изменен состав."
    )
         # Проверка Контрагенты
    xml_file_path_31 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Контрагенты.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_31,
        keyword_module,
        "✓ В Справочнике Контрагенты  Изменен состав",
        "✗ В Справочнике Контрагенты  не Изменен состав."
    )
         # Проверка Контрагенты
    xml_file_path_32 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Контрагенты.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_32,
        keyword_module,
        "✓ В Справочнике Контрагенты  Изменен состав",
        "✗ В Справочнике Контрагенты  не Изменен состав."
    )
     # Проверка Контрагенты
    xml_file_path_33 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Контрагенты\Forms\ФормаЭлемента\Ext\Form\Module.bsl"
    keyword_module = "ПРАКТИКОН"
    check_keyword_in_file(
        xml_file_path_33,
        keyword_module,
        "✓ В Справочнике Контрагенты  Изменена форма элемента",
        "✗ В Справочнике Контрагенты  не Изменена форма элемента."
    )
     # Проверка НаборыДополнительныхРеквизитовИСведений
    xml_file_path_34 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\НаборыДополнительныхРеквизитовИСведений.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_34,
        keyword_module,
        "✓ В Справочнике НаборыДополнительныхРеквизитовИСведений  Изменен состав",
        "✗ В Справочнике НаборыДополнительныхРеквизитовИСведений  не Изменен состав."
    )
    # Проверка Номенклатура (приведен к типовому)
    xml_file_path_35 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Номенклатура.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_35,
        keyword_module,
        "✓ В Справочнике Номенклатура  не Изменен состав",
        "✗ В Справочнике Номенклатура  Изменен состав.",
         invert_logic=True
    )
    # Проверка Организации
    xml_file_path_36 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Организации.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_36,
        keyword_module,
        "✓ В Справочнике Организации  Изменен состав",
        "✗ В Справочнике Организации  не Изменен состав."
    )
    # Проверка Партнеры
    xml_file_path_37 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Партнеры.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_37,
        keyword_module,
        "✓ В Справочнике Партнеры  Изменен состав",
        "✗ В Справочнике Партнеры  не Изменен состав."
    )
    # Проверка Пользователи
    xml_file_path_38 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Пользователи.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_38,
        keyword_module,
        "✓ В Справочнике Пользователи  Изменен состав",
        "✗ В Справочнике Пользователи  не Изменен состав."
    )
    # Проверка Пользователи
    xml_file_path_39 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Пользователи\Forms\ФормаЭлемента\Ext\Form\Module.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_39,
        keyword_module,
        "✓ В Справочнике Пользователи  Изменена форма элемента",
        "✗ В Справочнике Пользователи  не Изменена форма элемента.",
        required_count=8
    )
     # Проверка РолиИсполнителей
    xml_file_path_40 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\РолиИсполнителей.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_40,
        keyword_module,
        "✓ В Справочнике РолиИсполнителей  Изменен состав",
        "✗ В Справочнике РолиИсполнителей  не Изменен состав."
    )
      # Проверка СегментыНоменклатуры
    xml_file_path_41 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\СегментыНоменклатуры.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_41,
        keyword_module,
        "✓ В Справочнике СегментыНоменклатуры  Изменен состав",
        "✗ В Справочнике СегментыНоменклатуры  не Изменен состав."
    )
      # Проверка СегментыПартнеров
    xml_file_path_42 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\СегментыПартнеров.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_42,
        keyword_module,
        "✓ В Справочнике СегментыПартнеров  Изменен состав",
        "✗ В Справочнике СегментыПартнеров  не Изменен состав."
    )
     # Проверка СкидкиНаценки
    xml_file_path_43 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\СкидкиНаценки\Forms\ФормаЭлемента\Ext\Form\Module.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_43,
        keyword_module,
        "✓ В Справочнике СкидкиНаценки  Изменена форма элемента",
        "✗ В Справочнике СкидкиНаценки  не Изменена форма элемента.",
        required_count=20
    )
      # Проверка Склады
    xml_file_path_44 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\Склады.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_44,
        keyword_module,
        "✓ В Справочнике Склады  Изменен состав",
        "✗ В Справочнике Склады  не Изменен состав."
    )
      # Проверка СоглашенияСКлиентами
    xml_file_path_45 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\СоглашенияСКлиентами.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_45,
        keyword_module,
        "✓ В Справочнике СоглашенияСКлиентами  Изменен состав",
        "✗ В Справочнике СоглашенияСКлиентами  не Изменен состав."
    )
      # Проверка СоглашенияСПоставщиками
    xml_file_path_46 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\СоглашенияСПоставщиками.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_46,
        keyword_module,
        "✓ В Справочнике СоглашенияСПоставщиками  Изменен состав",
        "✗ В Справочнике СоглашенияСПоставщиками  не Изменен состав."
    )
     # Проверка СтруктураПредприятия
    xml_file_path_47 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\СтруктураПредприятия.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_47,
        keyword_module,
        "✓ В Справочнике СтруктураПредприятия  Изменен состав",
        "✗ В Справочнике СтруктураПредприятия  не Изменен состав."
    )
     # Проверка ТоварныеКатегории (приведен к типовому)
    xml_file_path_48 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ТоварныеКатегории.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_48,
        keyword_module,
        "✓ В Справочнике ТоварныеКатегории  не Изменен состав",
        "✗ В Справочнике ТоварныеКатегории  Изменен состав.",
         invert_logic=True
    )
 # Проверка ТранспортныеСредства
    xml_file_path_49 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ТранспортныеСредства\Ext\ObjectModule.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_49,
        keyword_module,
        "✓ В Справочнике ТранспортныеСредства  Изменен модуль объекта",
        "✗ В Справочнике ТранспортныеСредства  не Изменена модуль объекта.",
        required_count=4
    )
      # Проверка ТранспортныеСредства
    xml_file_path_50 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ТранспортныеСредства.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_50,
        keyword_module,
        "✓ В Справочнике ТранспортныеСредства  Изменен состав",
        "✗ В Справочнике ТранспортныеСредства  не Изменен состав."
    )
      # Проверка УсловияПредоставленияСкидокНаценок
    xml_file_path_51 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\УсловияПредоставленияСкидокНаценок.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_51,
        keyword_module,
        "✓ В Справочнике УсловияПредоставленияСкидокНаценок Добавлены предопределенные с префиксом АЗС_ ",
        "✗ В Скправочник УсловияПредоставленияСкидокНаценок не Добавлены предопределенные с префиксом АЗС_ ."
    )
      # Проверка ЦеновыеГруппы
    xml_file_path_52 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ЦеновыеГруппы.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_52,
        keyword_module,
        "✓ В Справочнике ЦеновыеГруппы Изменен состав ",
        "✗ В Справочнике ЦеновыеГруппы не Изменен состав."
    )
      # Проверка Внутреннее потребление
    xml_file_path_53 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ВнутреннееПотребление.xml"
    keyword_module = "AccumulationRegister.АЗС"
    check_keyword_in_file(
        xml_file_path_53,
        keyword_module,
        "✓ В Документ Внутреннее потребление добавлены движения АЗС ",
        "✗ В Документ Внутреннее потребление не добавлены движения АЗС ."
    )
      # Проверка ВозвратТоваровПоставщикуе
    xml_file_path_54 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ВозвратТоваровПоставщику.xml"
    keyword_module = "AccumulationRegister.АЗС"
    check_keyword_in_file(
        xml_file_path_54,
        keyword_module,
        "✓ В Документ ВозвратТоваровПоставщику добавлены движения АЗС ",
        "✗ В Документ ВозвратТоваровПоставщику не добавлены движения АЗС ."
    )
      # Проверка КорректировкаПриобретения
    xml_file_path_55 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаПриобретения.xml"
    keyword_module = "AccumulationRegister.АЗС"
    check_keyword_in_file(
        xml_file_path_55,
        keyword_module,
        "✓ В Документ КорректировкаПриобретения добавлены движения АЗС ",
        "✗ В Документ КорректировкаПриобретения не добавлены движения АЗС ."
    )
      # Проверка КорректировкаРеализации
    xml_file_path_56 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаРеализации.xml"
    keyword_module = "AccumulationRegister.АЗС"
    check_keyword_in_file(
        xml_file_path_56,
        keyword_module,
        "✓ В Документ КорректировкаРеализации добавлены движения АЗС ",
        "✗ В Документ КорректировкаРеализации не добавлены движения АЗС ."
    )
  
     # Проверка КорректировкаРегистров
    xml_file_path_57 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаРегистров.xml"
    keyword_module = "AccumulationRegister.АЗС"
    check_keyword_in_file(
        xml_file_path_57,
        keyword_module,
        "✓ В Документ КорректировкаРегистров добавлены движения АЗС ",
        "✗ В Документ КорректировкаРегистров не добавлены движения АЗС .",
        required_count=17
    )
      # Проверка ОприходованиеИзлишковТоваров
    xml_file_path_58 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ОприходованиеИзлишковТоваров.xml"
    keyword_module = "AccumulationRegister.АЗС"
    check_keyword_in_file(
        xml_file_path_58,
        keyword_module,
        "✓ В Документ ОприходованиеИзлишковТоваров добавлены движения АЗС ",
        "✗ В Документ ОприходованиеИзлишковТоваров не добавлены движения АЗС ."
    )
   # Проверка ПеремещениеТоваров
    xml_file_path_59 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПеремещениеТоваров.xml"
    keyword_module = "AccumulationRegister.АЗС"
    check_keyword_in_file(
        xml_file_path_59,
        keyword_module,
        "✓ В Документ ПеремещениеТоваров добавлены движения АЗС ",
        "✗ В Документ ПеремещениеТоваров не добавлены движения АЗС ."
    )
    # Проверка ПеремещениеТоваров
    xml_file_path_60 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПеремещениеТоваров\Ext\ObjectModule.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_60,
        keyword_module,
        "✓ В Документе ПеремещениеТоваров Изменен модуль объекта ",
        "✗ В Документе ПеремещениеТоваров не Изменен модуль объект а."
    )
    # Проверка ПересортицаТоваров
    xml_file_path_61 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПересортицаТоваров.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_61,
        keyword_module,
        "✓ В Документ ПересортицаТоваров Добавлены реквизиты АЗС ",
        "✗ В Документ ПересортицаТоваров не добавлены реквизиты АЗС ."
    )
     # Проверка ПересчетТоваров (приведен к типовому
    xml_file_path_62 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПересчетТоваров.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_62,
        keyword_module,
        "✓ В Документ ПересортицаТоваров не добавлены реквизиты АЗС ",
        "✗ В Документ ПересортицаТоваров добавлены реквизиты АЗС .",
        invert_logic=True
    )
     # Проверка ПланПродажПоКатегориям
    xml_file_path_63 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПланПродажПоКатегориям\Forms\ФормаДокумента\Ext\Form\Module.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_63,
        keyword_module,
        "✓ В Документ ПланПродажПоКатегориям добавлены изменения в модуль формы документа ",
        "✗ В Документ ПланПродажПоКатегориям не добавлены изменения в модуль формы документа .",
        required_count=36
    )
     # Проверка ПоступлениеБезналичныхДенежныхСредств
    xml_file_path_64 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПоступлениеБезналичныхДенежныхСредств\Forms\ФормаВыбора\Ext\Form\Module.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_64,
        keyword_module,
        "✓ В Документ ПоступлениеБезналичныхДенежныхСредств добавлены изменения в модуль формы выбора ",
        "✗ В Документ ПоступлениеБезналичныхДенежныхСредств не  изменения в модуль формы выбора."
    )
     # Проверка ПриходныйКассовыйОрдер
    xml_file_path_65 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПриходныйКассовыйОрдер.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_65,
        keyword_module,
        "✓ В Документ ПриходныйКассовыйОрдер добавлены реквизиты АЗС ",
        "✗ В Документ ПриходныйКассовыйОрдер не добавлены реквизиты АЗС ."
    )
     # Проверка РасходныйКассовыйОрдер
    xml_file_path_66 = r"E:\Users\AUshanov\izmenenie_tip\Documents\РасходныйКассовыйОрдер.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_66,
        keyword_module,
        "✓ В Документ РасходныйКассовыйОрдер добавлены реквизиты АЗС ",
        "✗ В Документ РасходныйКассовыйОрдер не добавлены реквизиты АЗС ."
    )
     
       # Проверка РеализацияТоваровУслуг
    xml_file_path_67 = r"E:\Users\AUshanov\izmenenie_tip\Documents\РеализацияТоваровУслуг.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_67,
        keyword_module,
        "✓ В Документ РасходныйКассовыйОрдер добавлены реквизиты АЗС ",
        "✗ В Документ РасходныйКассовыйОрдер не добавлены реквизиты АЗС ."
    )
       # Проверка РеализацияТоваровУслуг
    xml_file_path_68 = r"E:\Users\AUshanov\izmenenie_tip\Documents\РеализацияТоваровУслуг\Ext\ObjectModule.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_68,
        keyword_module,
        "✓ В Документ РеализацияТоваровУслуг  Изменен модуль объекта  ",
        "✗ В Документ РеализацияТоваровУслуг не Изменен модуль объекта  ."
    )
        # Проверка РегистрацияЦенНоменклатурыПоставщика
    xml_file_path_69 = r"E:\Users\AUshanov\izmenenie_tip\Documents\РегистрацияЦенНоменклатурыПоставщика.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_69,
        keyword_module,
        "✓ В Документе РегистрацияЦенНоменклатурыПоставщика добавлены реквизиты АЗС ",
        "✗ В Документе РегистрацияЦенНоменклатурыПоставщика не добавлены реквизиты АЗС ."
    )
            # Проверка СборкаТоваров
    xml_file_path_70 = r"E:\Users\AUshanov\izmenenie_tip\Documents\СборкаТоваров.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_70,
        keyword_module,
        "✓ В Документе СборкаТоваров добавлены реквизиты АЗС ",
        "✗ В Документе СборкаТоваров не добавлены реквизиты АЗС ."
    )
        # Проверка СписаниеНедостачТоваров
    xml_file_path_71 = r"E:\Users\AUshanov\izmenenie_tip\Documents\СписаниеНедостачТоваров.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_71,
        keyword_module,
        "✓ В Документе СписаниеНедостачТоваров добавлены реквизиты АЗС ",
        "✗ В Документе СписаниеНедостачТоваров не добавлены реквизиты АЗС ."
    )
        # Проверка УстановкаЦенНоменклатуры
    xml_file_path_72 = r"E:\Users\AUshanov\izmenenie_tip\Documents\РеализацияТоваровУслуг\Forms\ФормаДокумента\Ext\Form\Module.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_72,
        keyword_module,
        "✓ В Документе УстановкаЦенНоменклатуры добавлен код в  Модуль формы документа ",
        "✗ В Документе УстановкаЦенНоменклатуры не добавлены Модуль формы документ .",
        required_count=6
    )
     
        # Проверка УстановкаЦенНоменклатуры
    xml_file_path_72 = r"E:\Users\AUshanov\izmenenie_tip\Documents\УстановкаЦенНоменклатуры\Ext\ObjectModule.bsl"
    keyword_module = "Практикон"
    check_keyword_in_file(
        xml_file_path_72,
        keyword_module,
        "✓ В Документе УстановкаЦенНоменклатуры добавленн код в модуль объекта",
        "✗ В Документе УстановкаЦенНоменклатуры не добавлен код в модуль объекта .",
        required_count=6   
    )
        # Проверка ЧекККМ
    xml_file_path_73 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ЧекККМ.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_73,
        keyword_module,
        "✓ В Документе ЧекККМ добавлены движения АЗС ",
        "✗ В Документе ЧекККМ не добавлены Движения АЗС ."
    )
       # Проверка ЧекККМВозврат
    xml_file_path_74 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ЧекККМВозврат.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_74,
        keyword_module,
        "✓ В Документе ЧекККМВозврат добавлены движения АЗС ",
        "✗ В Документе ЧекККМВозврат не добавлены Движения АЗС ."
    )
       # Проверка СпособыПредоставленияСкидокНаценок
    xml_file_path_75 = r"E:\Users\AUshanov\izmenenie_tip\Enums\СпособыПредоставленияСкидокНаценок.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_75,
        keyword_module,
        "✓ В Перечислении СпособыПредоставленияСкидокНаценок добавлены Значение Перечисления АЗС ",
        "✗ В Перечислении СпособыПредоставленияСкидокНаценок не добавлены Значение Перечисления АЗС ."
    )
      # Проверка УсловияПредоставленияСкидокНаценок
    xml_file_path_76 = r"E:\Users\AUshanov\izmenenie_tip\Enums\УсловияПредоставленияСкидокНаценок.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_76,
        keyword_module,
        "✓ В Перечислении УсловияПредоставленияСкидокНаценок добавлены Значение Перечисления АЗС ",
        "✗ В Перечислении УсловияПредоставленияСкидокНаценок не добавлены Значение Перечисления АЗС ."
    )
      # Проверка ФормыОплаты
    xml_file_path_77 = r"E:\Users\AUshanov\izmenenie_tip\Enums\ФормыОплаты.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_77,
        keyword_module,
        "✓ В Перечислении ФормыОплаты добавлены Значение Перечисления АЗС ",
        "✗ В Перечислении ФормыОплаты не добавлены Значение Перечисления АЗС ."
    )
      # Проверка ХозяйственныеОперации
    xml_file_path_78 = r"E:\Users\AUshanov\izmenenie_tip\Enums\ХозяйственныеОперации.xml"
    keyword_module = "АЗС"
    check_keyword_in_file(
        xml_file_path_78,
        keyword_module,
        "✓ В Перечислении ХозяйственныеОперации добавлены Значение Перечисления АЗС ",
        "✗ В Перечислении ХозяйственныеОперации не добавлены Значение Перечисления АЗС ."
    )


    
    











#     # Путь к Оприходование излишков товаров
#     xml_file_path_24 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ОприходованиеИзлишковТоваров.xml"
#     keyword_2 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     check_keyword_in_file(
#         xml_file_path_24,
#         keyword_2,
#         "✓ ОприходованиеИзлишковТоваров Добавлен РН «АЗС_ Остатки Топлива».",
#         "✗ ОприходованиеИзлишковТоваров Не добавлен РН «АЗС_ Остатки Топлива»."
#     )
# #Путь к Оприходование излишков товаров
#     xml_file_path_3 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ОприходованиеИзлишковТоваров.xml"
#     keyword_3 = "<Name>УдалитьАЗС_ДокументОснование"
#     result_3 = check_keyword_in_file(
#         xml_file_path_3, 
#         keyword_3,
#         "✓ Реквизит АЗС_ДокументОснование переименован в УдалитьАЗС_ДокументОснование.",
#         "✗ Реквизит АЗС_ДокументОснование НЕ переименован в УдалитьАЗС_ДокументОснование."
#     )

# # Путь к Списание недостач товаров
#     xml_file_path_4 = r"E:\Users\AUshanov\izmenenie_tip\Documents\СписаниеНедостачТоваров.xml"
#     keyword_4 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_4 = check_keyword_in_file(
#         xml_file_path_4, 
#         keyword_4,
#         "✓ СписаниеНедостачТоваров Добавлен РН «АЗС_ Остатки Топлива».",
#         "✗ СписаниеНедостачТоваров Не добавлен РН «АЗС_ Остатки Топлива».")

# # Путь к Списание недостач товаров
#     xml_file_path_5 = r"E:\Users\AUshanov\izmenenie_tip\Documents\СписаниеНедостачТоваров.xml"
#     keyword_5 = "УдалитьАЗС_ДокументОснование"
#     result_5 = check_keyword_in_file(
#         xml_file_path_5, 
#         keyword_5,
#         "✓ СписаниеНедостачТоваров Есть УдалитьАЗС_ДокументОснование.",
#         "✗ СписаниеНедостачТоваров Нет УдалитьАЗС_ДокументОснование.")

# # Приобретение товаров услуг
#     xml_file_path_6 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПриобретениеТоваровУслуг.xml"
#     keyword_6 = "Удалить_АЗС_СебестоимостьТоваровОперативная"
#     result_6 = check_keyword_in_file(
#         xml_file_path_6, 
#         keyword_6,
#         "✗ ПриобретениеТоваровУслуг Есть Удалить_АЗС_СебестоимостьТоваровОперативная.",
#         "✓ ПриобретениеТоваровУслуг Нет Удалить_АЗС_СебестоимостьТоваровОперативная.")   

# # Корректировка регистров
#     xml_file_path_7 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаРегистров.xml"
#     keyword_7 = "Удалить_АЗС_СебестоимостьТоваровОперативная"
#     result_7 = check_keyword_in_file(
#         xml_file_path_7, 
#         keyword_7,
#         "✗ КорректировкаРегистров Есть Удалить_АЗС_СебестоимостьТоваровОперативная.",
#         "✓ КорректировкаРегистров Нет Удалить_АЗС_СебестоимостьТоваровОперативная.") 

# # Корректировка регистров
#     xml_file_path_8 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаРегистров.xml"
#     keyword_8 = "Удалить_АЗС_ВыручкаСебестоимостьПродажОперативная"
#     result_8 = check_keyword_in_file(
#         xml_file_path_8, 
#         keyword_8,
#         "✗ КорректировкаРегистров Есть Удалить_АЗС_ВыручкаСебестоимостьПродажОперативная.",
#         "✓ КорректировкаРегистров Нет Удалить_АЗС_ВыручкаСебестоимостьПродажОперативная.")
             
# # Корректировка регистров
#     xml_file_path_9 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаРегистров.xml"
#     keyword_9 = "Удалить_АЗС_РасчетыПоЭквайрингу"
#     result_9 = check_keyword_in_file(
#         xml_file_path_9, 
#         keyword_9,
#         "✗ КорректировкаРегистров Есть Удалить_АЗС_РасчетыПоЭквайрингу.",
#         "✓ КорректировкаРегистров Нет Удалить_АЗС_РасчетыПоЭквайрингу.")

# # КорректировкаПриобретения
#     xml_file_path_10 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаПриобретения.xml"
#     keyword_10 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_10 = check_keyword_in_file(
#         xml_file_path_10, 
#         keyword_10,
#         "✓ КорректировкаПриобретения Есть АЗС_ Остатки Топлива.",
#         "✗ КорректировкаПриобретения Нет АЗС_ Остатки Топлива.")

# # КорректировкаРеализации
#     xml_file_path_11 = r"E:\Users\AUshanov\izmenenie_tip\Documents\КорректировкаРеализации.xml"
#     keyword_11 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_11 = check_keyword_in_file(
#         xml_file_path_11, 
#         keyword_11,
#         "✓ КорректировкаРеализации Есть АЗС_ Остатки Топлива.",
#         "✗ КорректировкаРеализации Нет АЗС_ Остатки Топлива.")

# # ВнутреннееПотребление
#     xml_file_path_12 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ВнутреннееПотребление.xml"
#     keyword_12 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_12 = check_keyword_in_file(
#         xml_file_path_12, 
#         keyword_12,
#         "✓ ВнутреннееПотребление Есть АЗС_ Остатки Топлива.",
#         "✗ ВнутреннееПотребление Нет АЗС_ Остатки Топлива.")

# # Товарные категории
#     xml_file_path_13 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ТоварныеКатегории.xml"
#     keyword_13 = "УдалитьАЗС_КодВнешний"
#     result_13 = check_keyword_in_file(
#         xml_file_path_13, 
#         keyword_13,
#         "✗ ТоварныеКатегории есть УдалитьАЗС_КодВнешний.",
#         "✓ ТоварныеКатегории Нет УдалитьАЗС_КодВнешний.")      

# # Пересчет товаров
#     xml_file_path_14 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПересчетТоваров.xml"
#     keyword_14 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_14 = check_keyword_in_file(
#         xml_file_path_14,
#         keyword_14,
#         "✗ ПересчетТоваров есть АЗС_ОстаткиТоплива.",
#         "✓ ПересчетТоваров Нет АЗС_ОстаткиТоплива.")  

# # Приобретение товаров услуг
#     xml_file_path_15 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПриобретениеТоваровУслуг.xml"
#     keyword_15 = "<Name>УдалитьАЗС_ПоступлениеТоплива"
#     result_15 = check_keyword_in_file(
#         xml_file_path_15,
#         keyword_15,
#     "✓ ПриобретениеТоваровУслуг есть УдалитьАЗС_ПоступлениеТоплива.",
#     "✗ ПриобретениеТоваровУслуг Нет УдалитьАЗС_ПоступлениеТоплива.")

# # ПеремещениеТоваров
#     xml_file_path_16 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПеремещениеТоваров.xml"
#     keyword_16 = "<Name>УдалитьАЗС_ПоступлениеТоплива"
#     result_16 = check_keyword_in_file(
#         xml_file_path_16,
#         keyword_16,
#         "✓ ПеремещениеТоваров есть УдалитьАЗС_ПоступлениеТоплива.",
#         "✗ ПеремещениеТоваров Нет УдалитьАЗС_ПоступлениеТоплива.")


# # Приобретение товаров услуг
#     xml_file_path_17 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПриобретениеТоваровУслуг.xml"
#     keyword_17 = "<Name>АЗС_ДокументОснование"
#     result_17 = check_keyword_in_file(
#         xml_file_path_17, 
#         keyword_17,
#         "✓ ПриобретениеТоваровУслуг есть реквизит АЗС_ДокументОснование.",
#         "✗ ПриобретениеТоваровУслуг Нет реквизита АЗС_ДокументОснование.")


# # Путь к ПланПродажПоКатегориям
#     xml_file_path_18 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПланПродажПоКатегориям.xml"
#     keyword_18 = "<Name>АЗС_ПланироватьПоСумме"
#     result_18 = check_keyword_in_file(
#         xml_file_path_18, 
#         keyword_18,
#         "✗ ПланПродажПоКатегориям Есть АЗС_ПланироватьПоСумме.",
#         "✓ ПланПродажПоКатегориям Нет АЗС_ПланироватьПоСумме.")       

# # Путь к ПланПродажПоКатегориям
#     xml_file_path_19 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПланПродажПоКатегориям.xml"
#     keyword_19 = "<Name>АЗС_СуммаДокумента"
#     result_19 = check_keyword_in_file(
#         xml_file_path_19, 
#         keyword_19,
#         "✗ ПланПродажПоКатегориям Есть АЗС_СуммаДокумента.",
#         "✓ ПланПродажПоКатегориям Нет АЗС_СуммаДокумента.")

# # ПланПродажПоКатегориям.ФормаДокумента
# #   xml_file_path_2 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПланПродажПоКатегориям\Forms\ФормаДокумента\Ext\Form"
# #  keyword_2 = "<Name>АЗС_СуммаДокумента"
# #    result_2 = check_keyword_in_file(xml_file_path_2, keyword_2)
# #
# #    if result_2:
# #        print(Fore.Green + "✓ ПланПродажПоКатегориям.ФормаДокумента вынесен элемент формы Ответственный из группы АЗС_ГруппаИтоги.")
# #    else:
# #        print(Fore.RED + "✗ ПланПродажПоКатегориям.ФормаДокумента Не вынесен элемент формы Ответственный из группы АЗС_ГруппаИтоги.")

# # ПланПродажПоКатегориям.ФормаДокумента
#     xml_file_path_20 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПланПродажПоКатегориям\Forms\ФормаДокумента\Ext\Form\Module.bsl"
#     keyword_20 = "ПРАКТИКОН"
#     result_20 = check_keyword_in_file(
#         xml_file_path_20, 
#         keyword_20,
#         "✗ ПланПродажПоКатегориям.ФормаДокумента Есть доработки.",
#         "✓ ПланПродажПоКатегориям Нет Доработок.")

# # РНПланПродажПоКатегориям.ФормаДокумента
#     xml_file_path_21 = r"E:\Users\AUshanov\izmenenie_tip\AccumulationRegisters\ПланыПродажПоКатегориям.xml"
#     keyword_21 = "УдалитьАЗС_Сумма"
#     result_21 = check_keyword_in_file(
#         xml_file_path_21, 
#         keyword_21,
#         "✓ РНПланПродажПоКатегориям.ФормаДокумента Ресурс АЗС_Сумма переименован в УдалитьАЗС_Сумма.",
#         "✗ РНПланПродажПоКатегориям.ФормаДокумента  Ресурс АЗС_Сумма не переименован в УдалитьАЗС_Сумма.")       

# # ВариантыОтчетовПереопределяемый
#     xml_file_path_22 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ВариантыОтчетовПереопределяемый\Ext\Module.bsl"
#     keyword_22 = "АЗС_ВариантыОтчетовПереопределяемый.ОпределитьОбъектыСКомандамиОтчетов(Объекты)"
#     result_22 = check_keyword_in_file(
#         xml_file_path_22, 
#         keyword_22,
#         "✓ В ВариантыОтчетовПереопределяемый Есть строчка АЗС_ВариантыОтчетовПереопределяемый.ОпределитьОбъектыСКомандамиОтчетов(Объекты)",
#         "✗ в ВариантыОтчетовПереопределяемый НЕТ строчки АЗС_ВариантыОтчетовПереопределяемый.ОпределитьОбъектыСКомандамиОтчетов(Объекты)")


# # КритерийОтбораСвязанныеДокументы
#     xml_file_path_23 = r"E:\Users\AUshanov\izmenenie_tip\FilterCriteria\СвязанныеДокументы.xml"
#     keyword_23 = "Document.ПриобретениеТоваровУслуг.Attribute.АЗС_ДокументОснование"
#     result_23 = check_keyword_in_file(
#         xml_file_path_23, 
#         keyword_23,
#         "✓ В КритерийОтбораСвязанныеДокументы Добавлен в состав: ПриобретениеТоваровУслуг.Реквизиты.АЗС_ДокументОснование .",
#         "✗ В КритерийОтбораСвязанныеДокументы НЕ Добавлен в состав: ПриобретениеТоваровУслуг.Реквизиты.АЗС_ДокументОснование .")

#  # СозданиеНаОснованииПереопределяемый
#     xml_file_path_24 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\СозданиеНаОснованииПереопределяемый\Ext\Module.bsl"
#     keyword_24 = "АЗС_СозданиеНаОснованииПереопределяемый.ПриОпределенииОбъектовСКомандамиСозданияНаОсновании(Объекты)"
#     result_24 = check_keyword_in_file(
#         xml_file_path_24, 
#         keyword_24,
#         "✓ В СозданиеНаОснованииПереопределяемый Есть строчка АЗС_СозданиеНаОснованииПереопределяемый.ПриОпределенииОбъектовСКомандамиСозданияНаОсновании(Объекты)",
#         "✗ В СозданиеНаОснованииПереопределяемый НЕТ строчки АЗС_СозданиеНаОснованииПереопределяемый.ПриОпределенииОбъектовСКомандамиСозданияНаОсновании(Объекты)") 

#  # СозданиеНаОснованииПереопределяемый
#     xml_file_path_25 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\СозданиеНаОснованииПереопределяемый\Ext\Module.bsl"
#     keyword_25 = "АЗС_СозданиеНаОснованииПереопределяемый.ПриДобавленииКомандСозданияНаОсновании(Объект, КомандыСозданияНаОсновании"
#     result_25 = check_keyword_in_file(
#         xml_file_path_25, 
#         keyword_25,
#         "✓ В СозданиеНаОснованииПереопределяемый Есть строчка АЗС_СозданиеНаОснованииПереопределяемый.ПриДобавленииКомандСозданияНаОсновании(Объект, КомандыСозданияНаОсновании",
#         "✗ В СозданиеНаОснованииПереопределяемый НЕТ строчки АЗС_СозданиеНаОснованииПереопределяемый.ПриДобавленииКомандСозданияНаОсновании(Объект, КомандыСозданияНаОсновании")        

# # Путь к ПеремещениеТоваров
#     xml_file_path_26 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПеремещениеТоваров\Ext\ObjectModule.bsl"
#     keyword_26 = "ЗаполнитьПеремещениеТоваровПоПеремещениюТопливаИСопутствующихТоваров"
#     result_26 = check_keyword_in_file(
#         xml_file_path_26, 
#         keyword_26,
#         "✗ ПеремещениеТоваров В модуле объекта Не удалён код с комментариями Практикон",
#         "✓ ПеремещениеТоваров В модуле объекта удалён код с комментариями Практикон")    

#  # Приобретение товаров услуг
#     xml_file_path_27 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПриобретениеТоваровУслуг.xml"
#     keyword_27 = "<xr:Item xsi:type=""xr:MDObjectRef"">Document.АЗС_ПоступлениеТоплива</xr:Item>"
#     result_27 = check_keyword_in_file(
#         xml_file_path_27, 
#         keyword_27,
#         "✗ Приобретение товаров услуг В свойствах объекта Вводится на основании -> АЗС_ПоступлениеТоплива",
#         "✓ Приобретение товаров услуг В свойствах объекта нет Вводится на основании -> АЗС_ПоступлениеТоплива")                 
         
# # КритерийОтбораСвязанныеДокументы
#     xml_file_path_28 = r"E:\Users\AUshanov\izmenenie_tip\FilterCriteria\СвязанныеДокументы.xml"
#     keyword_28 = "Document.ПеремещениеТоваров.Attribute.АЗС_ДокументОснование"
#     result_28 = check_keyword_in_file(
#         xml_file_path_28, 
#         keyword_28,
#         "✓ В КритерийОтбораСвязанныеДокументы Добавлен в состав: ПеремещениеТоваров.Реквизиты.АЗС_ДокументОснование .",
#         "✗ В КритерийОтбораСвязанныеДокументы НЕ Добавлен в состав: ПеремещениеТоваров.Реквизиты.АЗС_ДокументОснование .") 

# # ЧекККМ
#     xml_file_path_29 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ЧекККМ.xml"
#     keyword_29 = "AccumulationRegister.АЗС_ПродажиКомплектовПоЧекам"
#     result_29 = check_keyword_in_file(
#         xml_file_path_29, 
#         keyword_29,
#         "✓ ЧекККМ Есть АЗС_ПродажиКомплектовПоЧекам.",
#         "✗ ЧекККМ Нет АЗС_ПродажиКомплектовПоЧекам.")

# # ЧекККМВозврат
#     xml_file_path_30 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ЧекККМВозврат.xml"
#     keyword_30 = "AccumulationRegister.АЗС_ПродажиКомплектовПоЧекам"
#     result_30 = check_keyword_in_file(
#         xml_file_path_30, 
#         keyword_30,
#         "✓ ЧекККМВозврат Есть АЗС_ПродажиКомплектовПоЧекам.",
#         "✗ ЧекККМВозврат Нет АЗС_ПродажиКомплектовПоЧекам.")

# # ПолучениеОбновленийПрограммыПереопределяемый
#     xml_file_path_31 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ПолучениеОбновленийПрограммыПереопределяемый\Ext\Module.bsl"
#     keyword_31 = "АЗС_ПолучениеОбновленийПрограммыПереопределяемый.ПриОпределенииНастроекЗагрузкиИсправлений(Настройки)"
#     result_31 = check_keyword_in_file(
#         xml_file_path_31, 
#         keyword_31,
#          "✓ В ПолучениеОбновленийПрограммыПереопределяемый Есть строчка АЗС_ПолучениеОбновленийПрограммыПереопределяемый.ПриОпределенииНастроекЗагрузкиИсправлений(Настройки)",
#          "✗ В ПолучениеОбновленийПрограммыПереопределяемый НЕТ АЗС_ПолучениеОбновленийПрограммыПереопределяемый.ПриОпределенииНастроекЗагрузкиИсправлений(Настройки)")        

# # ДоговорыКонтрагентов
#     xml_file_path_32 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ДоговорыКонтрагентов.xml"
#     keyword_32 = "<Name>АЗС_Госконтракт"
#     result_32 = check_keyword_in_file(
#         xml_file_path_32, 
#         keyword_32,
#         "✓ ЧекККМВозврат Есть реквизит АЗС_Госконтракт.",
#         "✗ ЧекККМВозврат Нет реквизита АЗС_Госконтракт.")

#  # СозданиеНаОснованииПереопределяемый
#     xml_file_path_33 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\СозданиеНаОснованииПереопределяемый\Ext\Module.bsl"
#     keyword_33 = "АЗС_СозданиеНаОснованииПереопределяемый.ПриДобавленииКомандСозданияНаОсновании(Объект, КомандыСозданияНаОсновании"
#     result_33 = check_keyword_in_file(
#         xml_file_path_33, 
#         keyword_33,
#         "✓ В СозданиеНаОснованииПереопределяемый Есть строчка АЗС_СозданиеНаОснованииПереопределяемый.ПриДобавленииКомандСозданияНаОсновании(Объект, КомандыСозданияНаОсновании",
#         "✗ В СозданиеНаОснованииПереопределяемый НЕТ АЗС_СозданиеНаОснованииПереопределяемый.ПриДобавленииКомандСозданияНаОсновании(Объект, КомандыСозданияНаОсновании")        
       
#  # ПересортицаТоваров
#     xml_file_path_34 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПересортицаТоваров.xml"
#     keyword_34 = "<Name>АЗС_ДокументОснование"
#     result_34 = check_keyword_in_file(
#         xml_file_path_34, 
#         keyword_34,
#         "✓ ПересортицаТоваров  Добавлен реквизит АЗС_ДокументОснование.",
#         "✗ ПересортицаТоваров НЕ добавлен реквизит АЗС_ДокументОснование.")      

# # КритерийОтбораСвязнненныеДокументы
#     xml_file_path_35 = r"E:\Users\AUshanov\izmenenie_tip\FilterCriteria\СвязанныеДокументы.xml"
#     keyword_35 = "Document.ПересортицаТоваров.Attribute.АЗС_ДокументОснование"
#     result_35 = check_keyword_in_file( 
#         xml_file_path_35, 
#         keyword_35,
#         "✓ В КритерийОтбораСвязанныеДокументы Добавлен в состав: ПересортицаТоваров.Реквизиты.АЗС_ДокументОснование.",
#         "✗ В КритерийОтбораСвязанныеДокументы НЕ Добавлен в состав: ПересортицаТоваров.Реквизиты.АЗС_ДокументОснование.")         

# # КритерийОтбораСвязнненныеДокументы
#     xml_file_path_36 = r"E:\Users\AUshanov\izmenenie_tip\FilterCriteria\СвязанныеДокументы.xml"
#     keyword_36 = "Document.ПересортицаТоваров.Attribute.АЗС_ПриобретениеТоваровУслуг"
#     result_36 = check_keyword_in_file(
#         xml_file_path_36,
#         keyword_36,
#         "✗ В КритерийОтбораСвязанныеДокументы НЕ Удалён из состава: ПересортицаТоваров.Реквизиты.АЗС_ПриобретениеТоваровУслуг (УдалитьАЗС_ПриобретениеТоваровУслуг.",
#         "✓ В КритерийОтбораСвязанныеДокументы Удалиён из состава: ПересортицаТоваров.Реквизиты.АЗС_ПриобретениеТоваровУслуг (УдалитьАЗС_ПриобретениеТоваровУслуг")

#  # ДоговорыКонтрагентов Реквизит АЗС_Госконтракт помечен как удаляемый.!!!!
#  #   xml_file_path_1 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ДоговорыКонтрагентов.xml"
#  #   keyword_1 = "<Name>АЗС_Госконтракт"
#  #   result_1 = check_keyword_in_file(xml_file_path_1, keyword_1)
#  #   if result_1:
#  #        print(Fore.GREEN + "✓ ЧекККМВозврат Есть реквизит АЗС_Госконтракт.")
#  #   else:
#  #       print(Fore.RED + "✗ ЧекККМВозврат Нет реквизита АЗС_Госконтракт.")  \             

# # ДоговорыКонтрагентов.ФормаЭлемента
#     xml_file_path_37 = r"E:\Users\AUshanov\izmenenie_tip\Catalogs\ДоговорыКонтрагентов\Forms\ФормаЭлемента\Ext\Form\Module.bsl"
#     keyword_37 = "Практикон"
#     result_37 = check_keyword_in_file(
#         xml_file_path_37, 
#         keyword_37,
#         "✗ ДоговорыКонтрагентов.ФормаЭлемент форма элемента НЕ ПРИВЕДЕНА к типовому виду.",
#         "✓ ДоговорыКонтрагентов.ФормаЭлемента Форма элемента приведена к типовому виду.")        

# # ТекущиеДелаПереопределяемый
#     xml_file_path_38 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ТекущиеДелаПереопределяемый\Ext\Module.bsl"
#     keyword_38 = "АЗС_ОбщегоНазначения.ПриОпределенииОбработчиковТекущихДел(ТекущиеДела)"
#     result_38 = check_keyword_in_file(
#         xml_file_path_38,
#         keyword_38,
#         "✓ В ТекущиеДелаПереопределяемый Есть строчка АЗС_ОбщегоНазначения.ПриОпределенииОбработчиковТекущихДел(ТекущиеДела)",
#         "✗ В ТекущиеДелаПереопределяемый НЕТ АЗС_ОбщегоНазначения.ПриОпределенииОбработчиковТекущихДел(ТекущиеДела)")        

# #УстановкаЦенНоменклатуры.ФормаДокумента
# #    module_file_path_39 = r"E:\Users\AUshanov\izmenenie_tip\Documents\УстановкаЦенНоменклатуры\Forms\ФормаДокумента\Ext\Form\Module.bsl"
# #keyword_39 = "ПРАКТИКОН"
# #
# #with open(module_file_path_39, "r", encoding="utf-8") as file_39:
# #    file_content_39 = file_39.read()
# #    count_keyword_39 = file_content_39.count(keyword_39)
# #
# #if count_keyword_39 == 2:
# #    print(Fore.GREEN + "✓ УстановкаЦенНоменклатуры.ФормаДокумента Найдено два слова 'ПРАКТИКОН'.")
# #elif count_keyword_39 == 3:
# #    print(Fore.RED + "✗ УстановкаЦенНоменклатуры.ФормаДокумента Найдено три слова 'ПРАКТИКОН'.")

       
# # ПользователиПереопределяемый
#     xml_file_path_40 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ПользователиПереопределяемый\Ext\Module.bsl"
#     keyword_40 = "АЗС_ПользователиПереопределяемый.ПриОпределенииНазначенияРолей(НазначениеРолей"
#     result_40 = check_keyword_in_file(
#         xml_file_path_40, 
#         keyword_40,
#         "✓ В ПользователиПереопределяемый Есть строчка  АЗС_ПользователиПереопределяемый.ПриОпределенииНазначенияРолей(НазначениеРолей)",
#         "✗ В ПользователиПереопределяемый НЕТ  АЗС_ПользователиПереопределяемый.ПриОпределенииНазначенияРолей(НазначениеРолей)")     

# #Командный интерфейс
#     xml_file_path_41 = r"E:\Users\AUshanov\izmenenie_tip\Ext\MainSectionCommandInterface.xml"
#     keyword_41 = "CommonCommand.АЗС_СводныйОтчетДляКлиента"
#     result_41 = check_keyword_in_file(
#         xml_file_path_41, 
#         keyword_41,
#         "✓ В Командном интерфейсе Добавлена общая команда (АЗС) Сводный отчет для клиента",
#         "✗ В Командном интерфейсе НЕ Добавлена общая команда (АЗС) Сводный отчет для клиента")   

# #Командный интерфейс
#     xml_file_path_42 = r"E:\Users\AUshanov\izmenenie_tip\Ext\MainSectionCommandInterface.xml"
#     keyword_42 = "CommonCommand.АЗС_ОтчетПоТопливнымКартамСамообслуживание"
#     result_42 = check_keyword_in_file(
#         xml_file_path_42, 
#         keyword_42,
#         "✓ В Командном интерфейсе Добавлена общая команда АЗС_ОтчетПоТопливнымКартамСамообслуживание",
#         "✗ В Командном интерфейсе НЕ Добавлена общая команда АЗС_ОтчетПоТопливнымКартамСамообслуживание")     

# # ОтгрузкаТоваровСХранения
#     xml_file_path_43 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ОтгрузкаТоваровСХранения.xml"
#     keyword_43 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_43 = check_keyword_in_file(
#         xml_file_path_43, 
#         keyword_43,
#         "✓ ОтгрузкаТоваровСХранения Есть РН АЗС_ Остатки Топлива.",
#         "✗ ОтгрузкаТоваровСХранения Нет РН АЗС_ Остатки Топлива.")

# # ПриемкаТоваровНаХранение 
#     xml_file_path_44 = r"E:\Users\E:\Users\AUshanov\izmenenie_tip\Documents\ПриемкаТоваровНаХранение.xml"
#     keyword_44 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_44 = check_keyword_in_file(
#         xml_file_path_44, 
#         keyword_44,
#         "✓ ПриемкаТоваровНаХранение  Есть РН АЗС_ Остатки Топлива.",
#         "✗ ПриемкаТоваровНаХранение  Нет РН АЗС_ Остатки Топлива.") 

# # ПередачаТоваровХранителю 
#     xml_file_path_45 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПередачаТоваровХранителю.xml"
#     keyword_45 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_45 = check_keyword_in_file(
#         xml_file_path_45, 
#         keyword_45,
#         "✓ ПередачаТоваровХранителю  Есть РН АЗС_ Остатки Топлива.",
#         "✗ ПередачаТоваровХранителю  Нет РН АЗС_ Остатки Топлива.")

# # ПоступлениеТоваровОтХранителя 
#     xml_file_path_46 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ПоступлениеТоваровОтХранителя.xml"
#     keyword_46 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_46 = check_keyword_in_file(
#         xml_file_path_46, 
#         keyword_46,
#         "✓ ПоступлениеТоваровОтХранителя  Есть РН АЗС_ Остатки Топлива.",
#         "✗ ПоступлениеТоваровОтХранителя  Нет РН АЗС_ Остатки Топлива.") 

# # ЧекККМ 
#     xml_file_path_47 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ЧекККМ.xml"
#     keyword_47 = "InformationRegister.АЗС_Талоны"
#     result_47 = check_keyword_in_file(
#         xml_file_path_47,
#         keyword_47,
#         "✓ ЧекККМ  Есть РС.АЗС_Талоны.",
#         "✗ ЧекККМ  Нет РС.АЗС_Талоны.")    

# # СчетНаОплатуКлиенту 
#     xml_file_path_48 = r"E:\Users\AUshanov\izmenenie_tip\Documents\СчетНаОплатуКлиенту.xml"
#     keyword_48 = "<Name>АЗС_ДокументОснование"
#     result_48 = check_keyword_in_file(
#         xml_file_path_48, 
#         keyword_48,
#         "✓ СчетНаОплатуКлиенту  Добавлен реквизит АЗС_ДокументОснование.",
#         "✗ СчетНаОплатуКлиенту  Не Добавлен реквизит АЗС_ДокументОснование.")               

# # КритерийОтбораСвязанныеДокументы
#     xml_file_path_49 = r"E:\Users\AUshanov\izmenenie_tip\FilterCriteria\СвязанныеДокументы.xml"
#     keyword_49 = "cfg:DocumentRef.АЗС_ЗаявкаНаПополнение"
#     result_49 = check_keyword_in_file(
#         xml_file_path_49, 
#         keyword_49,
#         "✓ В КритерийОтбораСвязанныеДокументы Изменен тип: добавлен ДокументСсылка.АЗС_ЗаявкаНаПополнение.",
#         "✗ В КритерийОтбораСвязанныеДокументы НЕ Изменен тип: добавлен ДокументСсылка.АЗС_ЗаявкаНаПополнение .") 

# # КритерийОтбораСвязанныеДокументы
#     xml_file_path_50 = r"E:\Users\AUshanov\izmenenie_tip\FilterCriteria\СвязанныеДокументы.xml"
#     keyword_50 = "Document.АЗС_ОперацииПоКартам.Attribute.ДокументОплаты"
#     result_50 = check_keyword_in_file(
#         xml_file_path_50, 
#         keyword_50,
#         "✓ В КритерийОтбораСвязанныеДокументы Изменен состав: добавлены реквизит Документ.АЗС_ОперацииПоКартам.Реквизит.ДокументОплаты.",
#         "✗ В КритерийОтбораСвязанныеДокументы НЕ добавлен реквизит Документ.АЗС_ОперацииПоКартам.Реквизит.ДокументОплаты.")  

# # КритерийОтбораСвязанныеДокументы
#     xml_file_path_50 = r"E:\Users\AUshanov\izmenenie_tip\FilterCriteria\СвязанныеДокументы.xml"
#     keyword_50 = "Document.СчетНаОплатуКлиенту.Attribute.АЗС_ДокументОснование"
#     result_50 = check_keyword_in_file(
#         xml_file_path_50, 
#         keyword_50,
#         "✓ В КритерийОтбораСвязанныеДокументы Изменен состав: добавлены реквизит Документ.СчетНаОплатуКлиенту.Реквизит.АЗС_ДокументОснование.",
#         "✗ В КритерийОтбораСвязанныеДокументы НЕ добавлен реквизит Документ.СчетНаОплатуКлиенту.Реквизит.АЗС_ДокументОснование.")

# # Путь к СтруктураПодчиненностиПереопределяемый
     
#     xml_file_path_51 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\СтруктураПодчиненностиПереопределяемый\Ext\Module.bsl"
#     keyword_module_51 = "АЗС_СтруктураПодчиненностиПереопределяемый.ПриПолученииПредставления(ТипДанных, Данные, Представление"
#     result_module_51 = check_keyword_in_file(
#         xml_file_path_51, 
#         keyword_module_51,
#         "✓ СтруктураПодчиненностиПереопределяемый Добавлен код АЗС_СтруктураПодчиненностиПереопределяемый.ПриПолученииПредставления(ТипДанных, Данные, Представление",
#         "✗ СтруктураПодчиненностиПереопределяемый не АЗС_СтруктураПодчиненностиПереопределяемый.ПриПолученииПредставления(ТипДанных, Данные, Представление")

# #Путь к Оприходование излишков товаров
#     xml_file_path_52 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ОприходованиеИзлишковТоваров.xml"
#     keyword_52 = "<Name>АЗС_ДокументОснование"
#     result_52 = check_keyword_in_file(
#         xml_file_path_52, 
#         keyword_52,
#             "✓ В Оприходование излишков товаров Добавлен реквизит АЗС_ДокументОснование.",
#             "✗ В Оприходование излишков товаров НЕ Добавлен реквизит АЗС_ДокументОснование.")

# #ВозвратТоваровОтКлиента
#     xml_file_path_53 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ВозвратТоваровОтКлиента.xml"
#     keyword_53 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_53 = check_keyword_in_file(
#         xml_file_path_53, 
#         keyword_53,
#         "✓ ВозвратТоваровОтКлиента Добавлен РН «АЗС_ Остатки Топлива».",
#         "✗ В ВозвратТоваровОтКлиента НЕ Добавлен РН «АЗС_ Остатки Топлива».")

# #ВозвратТоваровПоставщику
#     xml_file_path_54 = r"E:\Users\AUshanov\izmenenie_tip\Documents\ВозвратТоваровПоставщику.xml"
#     keyword_54 = "AccumulationRegister.АЗС_ОстаткиТоплива"
#     result_54 = check_keyword_in_file(
#         xml_file_path_54, 
#         keyword_54,
#         "✓ ВозвратТоваровПоставщику Добавлен РН «АЗС_ Остатки Топлива».",
#         "✗ В ВозвратТоваровПоставщику НЕ Добавлен РН «АЗС_ Остатки Топлива».")

# #МодульСеанса
#     xml_file_path_55 = r"E:\Users\AUshanov\izmenenie_tip\Ext\SessionModule.bsl"
#     keyword_55 = "АЗС_МодульСеанса.УстановкаПараметровСеанса(ИменаПараметровСеанса"
#     result_55 = check_keyword_in_file(
#         xml_file_path_55, 
#         keyword_55,
#         "✓ МодульСеанса Добавлен АЗС_МодульСеанса.УстановкаПараметровСеанса(ИменаПараметровСеанса).",
#         "✗ В МодульСеанса НЕ Добавлен РН АЗС_МодульСеанса.УстановкаПараметровСеанса(ИменаПараметровСеанса).")

# #ПечатьЭтикетокИЦенников
#     xml_file_path_56 = r"E:\Users\AUshanov\izmenenie_tip\DataProcessors\ПечатьЭтикетокИЦенников\Ext\ManagerModule.bsl"
#     keyword_56 = "АЗС_ОбщегоНазначения.АЗС_СвернутьТаблицуПередПечатью(СтруктураРезультата.Таблица)"
#     result_56 = check_keyword_in_file(
#         xml_file_path_56, 
#         keyword_56,
#         "✓ ПечатьЭтикетокИЦенниковДобавлен АЗС_ОбщегоНазначения.АЗС_СвернутьТаблицуПередПечатью(СтруктураРезультата.Таблица).",
#         "✗ В ПечатьЭтикетокИЦенников НЕТ АЗС_ОбщегоНазначения.АЗС_СвернутьТаблицуПередПечатью(СтруктураРезультата.Таблица).")

# #ДатыЗапретаИзмененияПереопределяемый
#     xml_file_path_57 = r"E:\Users\AUshanov\izmenenie_tip\CommonModules\ДатыЗапретаИзмененияПереопределяемый\Ext\Module.bsl"
#     keyword_57 = "АЗС_ОбщегоНазначения.ЗаполнитьИсточникиДанныхДляПроверкиЗапретаИзменения(ИсточникиДанных)"
#     result_57 = check_keyword_in_file(
#         xml_file_path_57, 
#         keyword_57,
#         "✓ ДатыЗапретаИзмененияПереопределяемый Есть АЗС_ОбщегоНазначения.ЗаполнитьИсточникиДанныхДляПроверкиЗапретаИзменения(ИсточникиДанных).",
#         "✗ В ДатыЗапретаИзмененияПереопределяемый НЕТ АЗС_ОбщегоНазначения.ЗаполнитьИсточникиДанныхДляПроверкиЗапретаИзменения(ИсточникиДанных).")
    
    # Запись значений из хранилища в Excel
    output_file_path = r"E:\Users\AUshanov\output.xlsx"
    
   # Проверка, существует ли файл Excel
if os.path.exists(output_file_path):
    # Если файл существует, то просто откроем его и очистим все данные
    workbook = load_workbook(output_file_path)
    sheet = workbook.active
    
    # Очистка всех данных на листе (удаление всех строк начиная с 2-й)
    sheet.delete_rows(2, sheet.max_row)  # Удаляем все строки, кроме первой (заголовков)
else:
    # Если файл не существует, создаем новый
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Статус", "Сообщение"])  # Заголовки столбцов

# Заполнение Excel файла
for status, message in results_storage:
    sheet.append([status, message])

# Сохранение изменений в файл
workbook.save(output_file_path)
     # Ожидание, чтобы консоль не закрылась
input("Нажмите Enter для выхода...")
